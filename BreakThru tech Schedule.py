#Mathew Jun, Sriya Peddinti, Sharanya Padithaya
#Automatic Schedule
#January 18, 2023


#using openPyxl to open Excel file
import openpyxl

#Input excel spread sheet into the green quotes:a
availability = openpyxl.load_workbook('Availability.xlsx')
availability_sheet = availability.active

#Create a new schedule spreadsheet
schedule = openpyxl.Workbook()
Excel = schedule.active
Excel.title = "Schedule"

#header row for the schedule
Excel['A1'] = "Monday"
Excel['B1'] = "Tuesday"
Excel['C1'] = "Wednesday"
Excel['D1'] = "Thursday"
Excel['E1'] = "Friday"

#Iterate with for loop (repeat over for each name) through the workdays spreadsheet 
#and write the schedule

#for each column with the headers name-friday, be able to read each cell
for row in range(2, availability_sheet.max_row + 1):
    name = availability_sheet.cell(row=row, column=1).value
    monday = availability_sheet.cell(row=row, column=2).value
    tuesday = availability_sheet.cell(row=row, column=3).value
    wednesday = availability_sheet.cell(row=row, column=4).value
    thursday = availability_sheet.cell(row=row, column=5).value
    friday = availability_sheet.cell(row=row, column=6).value
    
#if user types Y on excel sheet it will print their name on the schedule, 
#if nothing it will not print name
    Excel.cell(row=row, column=1).value = name if monday == "Y" else ""
    Excel.cell(row=row, column=2).value = name if tuesday == "Y" else ""
    Excel.cell(row=row, column=3).value = name if wednesday == "Y" else ""
    Excel.cell(row=row, column=4).value = name if thursday == "Y" else ""
    Excel.cell(row=row, column=5).value = name if friday == "Y" else ""

#Save the schedule spreadsheet
#Insert a name of what you want the spread sheet name to be 
schedule.save("schedule.xlsx")