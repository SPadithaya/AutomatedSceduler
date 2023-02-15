import openpyxl
from openpyxl import Workbook

dataframe = openpyxl.load_workbook('Schedule.xlsx')
dataframe1 = dataframe.active

#Creating a sheet within the workbook (this is the output excel file)
wb = Workbook()
wb.save("FinalSchedule.xlsx")
sheet1 = wb.create_sheet(index = 0, title = "Week Schedule")
wb.save("FinalSchedule.xlsx")

#Writing the Time Slots in the Ouput Excel Sheet
sheet1.cell(1, 1).value = "Times"
sheet1.cell(1, 2).value = "Person/# of People on Shift"
sheet1.cell(2, 1).value = "9:00"
sheet1.cell(3, 1).value = "10:00"
sheet1.cell(4, 1).value = "11:00"
sheet1.cell(5, 1).value = "12:00"
sheet1.cell(6, 1).value = "1:00"
sheet1.cell(7, 1).value = "2:00"
sheet1.cell(8, 1).value = "3:00"
sheet1.cell(9, 1).value = "4:00"
sheet1.cell(10, 1).value = "5:00"

#variables that count if there are overlaps in available times 
count9 = 1
count10 = 1
count11 = 1
count12 = 1
count1 = 1
count2 = 1
count3 = 1
count4 = 1
count5 = 1

#Reads input excel file, and writes name of person with that available time
#This code does not check for overlaps in available times
for i in range(2, dataframe1.max_row + 1):
  for j in range(3, dataframe1.max_column + 1):
   
    stg = str(dataframe1.cell(i, j).value)
    list = stg.split("-")
    row_num = 2

   
    for x in list:
      row_num = 2
      #checking 9:00am
      if sheet1.cell(row_num, 2).value is None:    
        if x == "9":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value
      else:
        if x == "9":
            count9 += 1
            sheet1.cell(row_num, 2).value = count9
      row_num += 1
      #checking 10:00am
      if sheet1.cell(row_num, 2).value is None:
        if x == "10":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value
      else:
        if x == "10":
            count10 += 1
            sheet1.cell(row_num, 2).value = count10
      row_num += 1
      #checking 11:00am
      if sheet1.cell(row_num, 2).value is None:
        if x == "11":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value
      else:
        if x == "11":
            count11 += 1
            sheet1.cell(row_num, 2).value = count11
      row_num += 1
      #Checking 12:00am
      if sheet1.cell(row_num, 2).value is None:
        if x == "12":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value
      else:
        if x == "12":
            count12 += 1
            sheet1.cell(row_num, 2).value = count12
      row_num += 1
      #checking 1:00pm
      if sheet1.cell(row_num, 2).value is None:
        if x == "1":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value
      else:
        if x == "1":
            count1 += 1
            sheet1.cell(row_num, 2).value = count1
      row_num += 1
      #checking 2:00pm
      if sheet1.cell(row_num, 2).value is None:
        if x == "2":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value
      else:
        if x == "2":
            count2 += 1
            sheet1.cell(row_num, 2).value = count2
      row_num += 1
      #checking 3:00pm
      if sheet1.cell(row_num, 2).value is None:
        if x == "3":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value        
      else:
        if x == "3":
            count3 += 1
            sheet1.cell(row_num, 2).value = count3
      row_num += 1
      #checking 4:00pm
      if sheet1.cell(row_num, 2).value is None:
        if x == "4":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value 
      else:
        if x == "4":
            count4 += 1
            sheet1.cell(row_num, 2).value = count4
      row_num += 1
      #checking 5:00pm
      if sheet1.cell(row_num, 2).value is None:
        if x == "5":
            sheet1.cell(row_num, 2).value = dataframe1.cell(i, 1).value  
      else:
        if x == "5":
            count5 += 1
            sheet1.cell(row_num, 2).value = count5
wb.save('FinalSchedule.xlsx')

